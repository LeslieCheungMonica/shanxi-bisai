import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: python read_xlsx.py <file.xlsx>")
    sys.exit(1)

wb = load_workbook(sys.argv[1], data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    for row in ws.iter_rows(values_only=True):
        print("\t".join(str(cell) if cell is not None else "" for cell in row))
    print()
